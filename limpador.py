# app_cafeteria_mestre.py
# Streamlit app para:
# 1) Padronizar o Excel quinzenal (PDC Campinas)
# 2) Gerar o relatório final (.xlsx) no padrão (A1/B1 + header na linha 2 + Tabela)
# 3) Atualizar um ARQUIVO MESTRE (append + dedupe SEM comer linhas + ordenação por data/hora desc)
#
# Requisitos:
#   pip install streamlit pandas openpyxl
# Execução:
#   streamlit run app_cafeteria_mestre.py

import streamlit as st
import pandas as pd
import re
from io import BytesIO
from datetime import datetime, date, time, timedelta

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


st.set_page_config(page_title="Cafeteria Access Report - PDC Campinas", layout="wide")
st.title("Cafeteria Access Report – PDC Campinas")


# ----------------------------
# Configurações
# ----------------------------
EXPECTED_COLUMNS = [
    "Date", "Time", "Event", "Reader", "SAPID", "LASTNAME", "FIRSTNAME",
    "MIDNAME", "BADGE", "Badge Type", "Company", "Unit Code", "Location"
]

SAPDC_KEYWORDS = [
    "SAPDC CAMPINAS",
    "SAPDC - CAMPINAS",
    "SAPDC-CAMPINAS",
    "PDC CAMPINAS",
    "SA PARTS DISTRIBUTION CENTER",
]

FILTER_COLUMNS = ["Reader", "Company", "Location", "Unit Code", "FIRSTNAME"]

MONTH_MAP = {
    "JAN": 1, "FEV": 2, "MAR": 3, "ABR": 4,
    "MAI": 5, "JUN": 6, "JUL": 7, "AGO": 8,
    "SET": 9, "OUT": 10, "NOV": 11, "DEZ": 12,
}


# ----------------------------
# Utilitários
# ----------------------------

def normalize_text(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza textos para facilitar filtro (sem transformar NaN em 'NAN')."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].fillna("").astype(str).str.upper().str.strip()
    return df


def enforce_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Garante todas as colunas esperadas na ordem certa."""
    df = df.copy()
    for col in EXPECTED_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[EXPECTED_COLUMNS]


def is_sapdc_campinas(row) -> bool:
    """Retorna True se qualquer coluna de filtro contém palavras-chave de SAPDC Campinas."""
    for col in FILTER_COLUMNS:
        if col in row.index:
            val = str(row[col])
            if any(k in val for k in SAPDC_KEYWORDS):
                return True
    return False


def detect_header_row(raw: pd.DataFrame) -> int | None:
    """Procura uma linha que contenha DATE e TIME (mais robusto)."""
    for i, row in raw.iterrows():
        cells = row.fillna("").astype(str).str.upper().str.strip().tolist()
        has_date = any((c == "DATE") or ("DATE" in c) for c in cells)
        has_time = any((c == "TIME") or ("TIME" in c) for c in cells)
        if has_date and has_time:
            return i
    # fallback: só DATE
    for i, row in raw.iterrows():
        cells = row.fillna("").astype(str).str.upper().str.strip().tolist()
        if any((c == "DATE") or ("DATE" in c) for c in cells):
            return i
    return None


def read_excel_auto(file_bytes: bytes) -> pd.DataFrame:
    """Lê Excel de entrada tentando:
    1) já padronizado (header na linha 2)
    2) header na primeira linha
    3) bruto (detectar linha do cabeçalho)
    """
    # tentativa 1: header na linha 2
    try:
        df = pd.read_excel(BytesIO(file_bytes), header=1, engine="openpyxl")
        hits = sum(1 for c in EXPECTED_COLUMNS if c in set(map(str, df.columns)))
        if hits >= 7:
            return df
    except Exception:
        pass

    # tentativa 2: header na linha 1
    try:
        df = pd.read_excel(BytesIO(file_bytes), header=0, engine="openpyxl")
        hits = sum(1 for c in EXPECTED_COLUMNS if c in set(map(str, df.columns)))
        if hits >= 7:
            return df
    except Exception:
        pass

    # tentativa 3: bruto
    raw = pd.read_excel(BytesIO(file_bytes), header=None, engine="openpyxl")
    header_row = detect_header_row(raw)
    if header_row is None:
        raise ValueError("Cabeçalho não encontrado (não achei DATE/TIME).")
    df = pd.read_excel(BytesIO(file_bytes), header=header_row, engine="openpyxl")
    return df


def extract_dates_from_filename(filename: str, year_default: int) -> tuple[str | None, str | None, str | None]:
    """Extrai datas do nome do arquivo.

    Aceita:
      - 16A22FEV2026
      - 24.11 a 30.11 (usa year_default se não houver ano)
      - 24.11 a 30.11.2026
    """
    name = filename.upper()

    m = re.search(r"(\d{2})A(\d{2})([A-Z]{3})(\d{4})", name)
    if m:
        d_start, d_end, mon, year = m.groups()
        month = MONTH_MAP.get(mon)
        if not month:
            return None, None, None
        start = f"{month}/{int(d_start)}/{year} 12:00:00 AM"
        end = f"{month}/{int(d_end)}/{year} 11:59:59 PM"
        label = f"{int(d_start):02d}.{month:02d} a {int(d_end):02d}.{month:02d}"
        return start, end, label

    m = re.search(r"(\d{2})\.(\d{2})\s*A\s*(\d{2})\.(\d{2})(?:\.(\d{4}))?", name)
    if m:
        d1, m1, d2, m2, y = m.groups()
        year = int(y) if y else year_default
        start = f"{int(m1)}/{int(d1)}/{year} 12:00:00 AM"
        end = f"{int(m2)}/{int(d2)}/{year} 11:59:59 PM"
        label = f"{int(d1):02d}.{int(m1):02d} a {int(d2):02d}.{int(m2):02d}"
        return start, end, label

    return None, None, None


def build_query(start: str, end: str) -> str:
    """Monta o texto da QUERY. Ajuste os READERS conforme necessário."""
    return (
        f"QUERY:  START DATE:  {start};   "
        f"END DATE:  {end};   "
        "READERS:  BR-GS-CATALAO CAFE 1 ENTRY, "
        "BR-GS-CATALAO CAFE 2 ENTRY, "
        "BR-SP-INTBA C&F 1 OFF CAFE TS #1 ENTRY, "
        "BR-SP-INTBA C&F 1 OFF CAFE TS #2 ENTRY, "
        "BR-SP-INTBA C&F 2 OFF CAFE TS..."
    )


# ----------------------------
# Parsing robusto de Date/Time
# ----------------------------

def _parse_date_series(s: pd.Series) -> pd.Series:
    """Converte para datetime.date (aceita string, datetime, excel serial)."""
    if s is None:
        return s
    ss = s.copy()

    # Excel serial (número)
    if pd.api.types.is_numeric_dtype(ss):
        dt = pd.to_datetime(ss, errors="coerce", unit="D", origin="1899-12-30")
        return dt.dt.date

    # datetime64
    if pd.api.types.is_datetime64_any_dtype(ss):
        return ss.dt.date

    # objetos: pode ter date/datetime
    def conv(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return pd.NaT
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date) and not isinstance(v, datetime):
            return v
        try:
            return pd.to_datetime(v, errors="coerce").date()
        except Exception:
            return pd.NaT

    return ss.apply(conv)


def _excel_time_float_to_time(x):
    """Converte float excel (fração do dia) em datetime.time."""
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return pd.NaT
    try:
        seconds = int(round(float(x) * 86400))
        seconds = seconds % 86400
        h = seconds // 3600
        m = (seconds % 3600) // 60
        s = seconds % 60
        return time(h, m, s)
    except Exception:
        return pd.NaT


def _parse_time_series(s: pd.Series) -> pd.Series:
    """Converte para datetime.time (aceita time, datetime, string, excel float)."""
    if s is None:
        return s
    ss = s.copy()

    # se já é datetime64 -> pega só hora
    if pd.api.types.is_datetime64_any_dtype(ss):
        return ss.dt.time

    # excel float
    if pd.api.types.is_numeric_dtype(ss):
        return ss.apply(_excel_time_float_to_time)

    # objetos (time/datetime/string)
    def conv(v):
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return pd.NaT
        if isinstance(v, time) and not isinstance(v, datetime):
            return v
        if isinstance(v, datetime):
            return v.time()
        # tenta string
        try:
            dt = pd.to_datetime(str(v), errors="coerce")
            if pd.isna(dt):
                # tenta formatos comuns
                for fmt in ("%H:%M", "%H:%M:%S"):
                    try:
                        return datetime.strptime(str(v).strip(), fmt).time()
                    except Exception:
                        pass
                return pd.NaT
            return dt.time()
        except Exception:
            return pd.NaT

    return ss.apply(conv)


def parse_datetime_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Garante Date e Time consistentes sem sumir com a hora das listas antigas."""
    df = df.copy()

    if "Date" in df.columns:
        df["Date"] = _parse_date_series(df["Date"])

    if "Time" in df.columns:
        df["Time"] = _parse_time_series(df["Time"])

    return df


def dedupe_sort(df: pd.DataFrame) -> pd.DataFrame:
    """Remove APENAS duplicados idênticos e ordena por Date/Time desc.

    Para não "comer" informação, a deduplicação é por TODAS as colunas esperadas.
    Assim só remove linhas exatamente iguais (duplicadas de verdade).
    """
    df = df.copy()

    # dedupe seguro: só remove duplicata perfeita
    df = df.drop_duplicates(subset=[c for c in EXPECTED_COLUMNS if c in df.columns], keep="last")

    # ordenação: combina Date + Time em um datetime (quando possível)
    def to_dt(d, t):
        try:
            if pd.isna(d):
                return pd.NaT
            if pd.isna(t):
                return datetime.combine(d, time(0, 0, 0))
            return datetime.combine(d, t)
        except Exception:
            return pd.NaT

    dcol = df["Date"] if "Date" in df.columns else pd.Series([pd.NaT] * len(df))
    tcol = df["Time"] if "Time" in df.columns else pd.Series([pd.NaT] * len(df))

    df["__dt_sort"] = [to_dt(d, t) for d, t in zip(dcol, tcol)]
    df = df.sort_values(by="__dt_sort", ascending=False, na_position="last")
    df = df.drop(columns=["__dt_sort"], errors="ignore")

    return df


# ----------------------------
# Excel writer
# ----------------------------

def make_excel_bytes(df_final: pd.DataFrame, query: str, title: str = "Cafeteria Access Report") -> bytes:
    """Cria um Excel no padrão (A1/B1 + header na linha 2 + tabela) e aplica formatos.

    IMPORTANTE: garante que a coluna Time seja salva como horário (hh:mm:ss)
    para não sumir no Excel quando juntar períodos.
    """
    # garante parse antes de escrever
    df_final = parse_datetime_cols(df_final)

    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_final.to_excel(writer, index=False, startrow=1, sheet_name="Report")

    out.seek(0)
    wb = load_workbook(out)
    ws = wb.active

    ws["A1"] = title
    ws["B1"] = query

    # Formatação de Date/Time (pra não aparecer em branco)
    headers = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
    header_to_col = {h: idx + 1 for idx, h in enumerate(headers) if h is not None}

    date_col = header_to_col.get("Date")
    time_col = header_to_col.get("Time")

    if date_col:
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=date_col)
            if cell.value is not None:
                cell.number_format = "dd/mm/yyyy"

    if time_col:
        for r in range(3, ws.max_row + 1):
            cell = ws.cell(row=r, column=time_col)
            if cell.value is not None:
                cell.number_format = "hh:mm:ss"

    last_row = ws.max_row
    last_col = ws.max_column
    table_ref = f"A2:{get_column_letter(last_col)}{last_row}"

    # remove tabelas existentes (se houver)
    if getattr(ws, "tables", None):
        for tname in list(ws.tables.keys()):
            del ws.tables[tname]

    table = Table(displayName="CafeteriaAccessReport", ref=table_ref)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)

    ws.freeze_panes = "A3"

    out2 = BytesIO()
    wb.save(out2)
    return out2.getvalue()


def read_master_df(master_bytes: bytes) -> tuple[pd.DataFrame, str, str]:
    """Lê o arquivo mestre e devolve (df, title, query)."""
    wb = load_workbook(BytesIO(master_bytes), data_only=True)
    ws = wb.active

    title = ws["A1"].value or "Cafeteria Access Report"
    query = ws["B1"].value or "MASTER FILE - Consolidado"

    df = pd.read_excel(BytesIO(master_bytes), header=1, engine="openpyxl")
    df = enforce_columns(df)
    df = parse_datetime_cols(df)
    return df, title, query


# ----------------------------
# Interface
# ----------------------------

st.subheader("1) Gerar relatório quinzenal (padrão)")

apply_filter = st.checkbox("Aplicar filtro SAPDC Campinas", value=True)
year_default = st.number_input("Ano (usado se o nome do arquivo não tiver ano)", value=datetime.now().year, step=1)

uploaded_file = st.file_uploader("Upload do arquivo Excel (.xlsx)", type=["xlsx"], key="up_quinz")

if uploaded_file:
    file_bytes = uploaded_file.getvalue()
    filename = uploaded_file.name

    df_raw = read_excel_auto(file_bytes)
    df_norm = normalize_text(df_raw)

    if apply_filter:
        mask = df_norm.apply(is_sapdc_campinas, axis=1)
        df_filtered = df_raw.loc[mask].copy()
    else:
        df_filtered = df_raw.copy()

    if df_filtered.empty:
        st.error("Nenhum registro encontrado (após filtro).")
        st.stop()

    df_final = enforce_columns(df_filtered)
    df_final = parse_datetime_cols(df_final)

    # datas/query/nome
    start_date, end_date, label = extract_dates_from_filename(filename, year_default)
    if start_date:
        query = build_query(start_date, end_date)
        output_name = f"Relatorio de Controle do Restaurante {label}.xlsx"
    else:
        query = ""
        output_name = "Relatorio de Controle do Restaurante Padronizado.xlsx"

    st.metric("Total de itens no relatório", len(df_final))
    st.dataframe(df_final, use_container_width=True)

    report_bytes = make_excel_bytes(df_final, query)

    st.download_button(
        "Baixar relatório final",
        data=report_bytes,
        file_name=output_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.divider()
    st.subheader("2) Atualizar ARQUIVO MESTRE (append + sem duplicar cabeçalho + ordenar por data/hora)")

    st.caption(
        "Você pode atualizar o mestre de 2 formas:\n"
        "• (A) Upload do mestre e baixar atualizado\n"
        "• (B) Salvar direto em um caminho local (Windows / pasta de rede)"
    )

    mode = st.radio(
        "Modo de atualização do mestre",
        options=["A) Upload do mestre e baixar atualizado", "B) Salvar em um caminho local (Windows)"],
        index=0,
    )

    if mode.startswith("A"):
        master_up = st.file_uploader("Upload do arquivo mestre (se já existir)", type=["xlsx"], key="up_master")

        if st.button("🔄 Atualizar arquivo mestre", type="primary"):
            if master_up:
                master_bytes = master_up.getvalue()
                df_master, title_master, query_master = read_master_df(master_bytes)
            else:
                df_master = pd.DataFrame(columns=EXPECTED_COLUMNS)
                title_master = "Cafeteria Access Report"
                query_master = "MASTER FILE - Consolidado"

            # une SEM comer linhas
            df_all = pd.concat([df_master, df_final], ignore_index=True)
            df_all = enforce_columns(df_all)
            df_all = parse_datetime_cols(df_all)
            df_all = dedupe_sort(df_all)

            master_updated_bytes = make_excel_bytes(df_all, query_master, title=title_master)

            st.success("Arquivo mestre atualizado! (horários preservados, sem cabeçalho duplicado e sem perda de linhas)")
            st.download_button(
                "⬇️ Baixar arquivo mestre atualizado",
                data=master_updated_bytes,
                file_name="Arquivo_Mestre_Cafeteria_PDC_Campinas.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    else:
        st.warning(
            "Esse modo só funciona se você estiver rodando o Streamlit localmente (no Windows), "
            "com permissão de escrita na pasta da rede/local."
        )

        master_path = st.text_input(
            "Caminho completo do arquivo mestre (ex.: \\\\servidor\\pasta\\Arquivo_Mestre.xlsx)",
            value=r"\\3qfiler\3q\DRHU\RHSA\HRSAPDC\Gestão De Refeição\Arquivo_Mestre_Cafeteria_PDC_Campinas.xlsx",
        )

        if st.button("🔄 Atualizar arquivo mestre", type="primary"):
            try:
                import os

                if os.path.exists(master_path):
                    with open(master_path, "rb") as f:
                        master_bytes = f.read()
                    df_master, title_master, query_master = read_master_df(master_bytes)
                else:
                    df_master = pd.DataFrame(columns=EXPECTED_COLUMNS)
                    title_master = "Cafeteria Access Report"
                    query_master = "MASTER FILE - Consolidado"

                df_all = pd.concat([df_master, df_final], ignore_index=True)
                df_all = enforce_columns(df_all)
                df_all = parse_datetime_cols(df_all)
                df_all = dedupe_sort(df_all)

                master_updated_bytes = make_excel_bytes(df_all, query_master, title=title_master)

                os.makedirs(os.path.dirname(master_path), exist_ok=True)
                with open(master_path, "wb") as f:
                    f.write(master_updated_bytes)

                st.success(f"Mestre atualizado e salvo em: {master_path}")

            except Exception as e:
                st.error(f"Falha ao salvar no caminho informado: {e}")
